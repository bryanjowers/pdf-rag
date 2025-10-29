#!/usr/bin/env python3
"""
xlsx.py - Excel/CSV processing handler with smart table chunking

Processes XLSX and CSV files with semantic-aware table boundary detection.
Implements the chunking strategy from PRD Section "XLSX Chunking (MVP Simplified)".
"""

import json
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import pandas as pd


def process_xlsx(
    xlsx_path: Path,
    output_dir: Path,
    config: Dict,
    batch_id: str
) -> Dict:
    """
    Process Excel or CSV file with smart table chunking.

    Args:
        xlsx_path: Path to input XLSX or CSV file
        output_dir: Base output directory
        config: Configuration dictionary
        batch_id: Unique batch identifier

    Returns:
        Processing result dictionary with metadata
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"File not found: {xlsx_path}")

    start_time = time.time()
    warnings = []

    # Prepare output directories
    markdown_dir = output_dir / "markdown"
    jsonl_dir = output_dir / "jsonl"
    markdown_dir.mkdir(parents=True, exist_ok=True)
    jsonl_dir.mkdir(parents=True, exist_ok=True)

    stem = xlsx_path.stem
    markdown_path = markdown_dir / f"{stem}.md"
    jsonl_path = jsonl_dir / f"{stem}.jsonl"

    try:
        print(f"   🔄 Processing Excel file: {xlsx_path.name}")

        # Route based on extension
        if xlsx_path.suffix.lower() == '.csv':
            markdown_content, chunks = _process_csv(xlsx_path, config, batch_id)
        else:
            markdown_content, chunks = _process_xlsx(xlsx_path, config, batch_id)

        char_count = len(markdown_content)

        # Write markdown
        markdown_path.write_text(markdown_content, encoding="utf-8")

        # Write JSONL
        with jsonl_path.open("w", encoding="utf-8") as f:
            for chunk in chunks:
                f.write(json.dumps(chunk, ensure_ascii=False) + "\n")

        duration_ms = int((time.time() - start_time) * 1000)

        print(f"   ✅ Excel processing complete")
        print(f"      Output: {char_count:,} chars")
        print(f"      Duration: {duration_ms/1000:.1f}s")
        print(f"      Chunks: {len(chunks)}")

        return {
            "success": True,
            "processor": "openpyxl",
            "markdown_path": markdown_path,
            "jsonl_path": jsonl_path,
            "processing_duration_ms": duration_ms,
            "page_count": None,  # Not applicable for XLSX
            "char_count": char_count,
            "estimated_tokens": len(markdown_content.split()),
            "chunk_count": len(chunks),
            "warnings": warnings,
            "error": None
        }

    except Exception as e:
        duration_ms = int((time.time() - start_time) * 1000)
        error_msg = f"Excel processing failed: {str(e)}"
        print(f"   ❌ {error_msg}")

        return {
            "success": False,
            "processor": None,
            "markdown_path": None,
            "jsonl_path": None,
            "processing_duration_ms": duration_ms,
            "char_count": 0,
            "estimated_tokens": 0,
            "chunk_count": 0,
            "warnings": warnings,
            "error": error_msg
        }


def _process_xlsx(xlsx_path: Path, config: Dict, batch_id: str) -> Tuple[str, List[Dict]]:
    """
    Process XLSX file with multiple sheets.

    Returns:
        Tuple of (markdown_content, jsonl_chunks)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)  # Evaluate formulas
    xlsx_config = config.get("xlsx", {})

    all_markdown = []
    all_chunks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip hidden sheets
        if xlsx_config.get("skip_hidden_sheets", True) and ws.sheet_state == 'hidden':
            print(f"      Skipping hidden sheet: {sheet_name}")
            continue

        print(f"      Processing sheet: {sheet_name} ({ws.max_row} rows × {ws.max_column} cols)")

        # Extract sheet data
        sheet_data = []
        for row in ws.iter_rows(values_only=True):
            sheet_data.append(list(row))

        if not sheet_data:
            continue

        # Detect table boundaries and chunk
        table_chunks = detect_table_boundaries(sheet_data, xlsx_config)

        # Convert each chunk to markdown
        for chunk_info in table_chunks:
            md, jsonl = _table_to_markdown_and_jsonl(
                chunk_info,
                sheet_name,
                xlsx_path,
                config,
                batch_id
            )
            all_markdown.append(md)
            all_chunks.extend(jsonl)

    wb.close()

    # Combine all markdown
    combined_markdown = "\n\n---\n\n".join(all_markdown)
    return combined_markdown, all_chunks


def _process_csv(csv_path: Path, config: Dict, batch_id: str) -> Tuple[str, List[Dict]]:
    """
    Process CSV file as single table.

    Returns:
        Tuple of (markdown_content, jsonl_chunks)
    """
    df = pd.read_csv(csv_path)
    xlsx_config = config.get("xlsx", {})

    # Convert to list of lists
    data = [df.columns.tolist()] + df.values.tolist()

    # Detect boundaries
    table_chunks = detect_table_boundaries(data, xlsx_config)

    # Convert to markdown
    all_markdown = []
    all_chunks = []

    for chunk_info in table_chunks:
        md, jsonl = _table_to_markdown_and_jsonl(
            chunk_info,
            "CSV",
            csv_path,
            config,
            batch_id
        )
        all_markdown.append(md)
        all_chunks.extend(jsonl)

    combined_markdown = "\n\n".join(all_markdown)
    return combined_markdown, all_chunks


def detect_table_boundaries(
    data: List[List],
    xlsx_config: Dict
) -> List[Dict]:
    """
    Detect logical table boundaries using heuristics from PRD.

    Rules:
    1. Split at blank rows (≥90% empty)
    2. Split when column schema changes (>30% difference)
    3. Detect mid-sheet headers (≥80% non-numeric cells)
    4. Hard cap at 2000 rows per chunk

    Args:
        data: 2D list of cell values
        xlsx_config: XLSX configuration from config

    Returns:
        List of chunk info dicts with table data
    """
    if not data:
        return []

    # Get config thresholds
    blank_threshold = xlsx_config.get("blank_row_threshold", 0.90)
    schema_threshold = xlsx_config.get("schema_change_threshold", 0.30)
    header_threshold = xlsx_config.get("header_detection_threshold", 0.80)
    max_rows = xlsx_config.get("max_rows_per_chunk", 2000)

    chunks = []
    current_chunk = []
    current_header = None
    last_schema = None

    for row_idx, row in enumerate(data):
        # Rule 1: Blank row detection
        non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())
        if non_empty / len(row) < (1 - blank_threshold) and current_chunk:
            # Blank row - finalize chunk
            chunks.append({
                "rows": [current_header] + current_chunk if current_header else current_chunk,
                "row_span": (row_idx - len(current_chunk), row_idx),
                "has_header": current_header is not None,
                "has_blank_boundary": True
            })
            current_chunk = []
            current_header = None
            last_schema = None
            continue

        # Rule 2: Column schema change detection
        current_schema = set(i for i, cell in enumerate(row) if cell is not None)
        if last_schema is not None:
            schema_diff = len(current_schema.symmetric_difference(last_schema)) / len(row)
            if schema_diff > schema_threshold and current_chunk:
                # Schema changed - finalize chunk
                chunks.append({
                    "rows": [current_header] + current_chunk if current_header else current_chunk,
                    "row_span": (row_idx - len(current_chunk), row_idx),
                    "has_header": current_header is not None,
                    "has_blank_boundary": False
                })
                current_chunk = []
                current_header = None

        last_schema = current_schema

        # Rule 3: Mid-sheet header detection
        non_numeric = sum(1 for cell in row if cell and not isinstance(cell, (int, float)))
        if non_numeric / non_empty >= header_threshold if non_empty > 0 else False:
            # Likely a header row
            if current_chunk:
                # Finalize previous chunk
                chunks.append({
                    "rows": [current_header] + current_chunk if current_header else current_chunk,
                    "row_span": (row_idx - len(current_chunk), row_idx),
                    "has_header": current_header is not None,
                    "has_blank_boundary": False
                })
                current_chunk = []
            current_header = row
            continue

        # Add to current chunk
        current_chunk.append(row)

        # Rule 4: Hard cap at max_rows
        if len(current_chunk) >= max_rows:
            chunks.append({
                "rows": [current_header] + current_chunk if current_header else current_chunk,
                "row_span": (row_idx - len(current_chunk) + 1, row_idx + 1),
                "has_header": current_header is not None,
                "has_blank_boundary": False
            })
            current_chunk = []
            # Keep header for next chunk

    # Add remaining chunk
    if current_chunk:
        chunks.append({
            "rows": [current_header] + current_chunk if current_header else current_chunk,
            "row_span": (len(data) - len(current_chunk), len(data)),
            "has_header": current_header is not None,
            "has_blank_boundary": False
        })

    return chunks


def _table_to_markdown_and_jsonl(
    chunk_info: Dict,
    sheet_name: str,
    source_path: Path,
    config: Dict,
    batch_id: str
) -> Tuple[str, List[Dict]]:
    """
    Convert table chunk to markdown and JSONL records.

    Args:
        chunk_info: Chunk info from detect_table_boundaries()
        sheet_name: Name of sheet
        source_path: Original file path
        config: Configuration
        batch_id: Batch ID

    Returns:
        Tuple of (markdown_text, jsonl_records)
    """
    from utils_classify import compute_file_hash, get_mime_type

    rows = chunk_info["rows"]
    if not rows:
        return "", []

    # Convert to markdown table
    markdown_lines = [f"## Sheet: {sheet_name}\n"]

    # Find max column count
    max_cols = max(len(row) for row in rows)

    # Render table
    for i, row in enumerate(rows):
        # Pad row to max_cols
        padded_row = list(row) + [None] * (max_cols - len(row))
        cells = [str(cell) if cell is not None else "" for cell in padded_row]
        markdown_lines.append("| " + " | ".join(cells) + " |")

        # Add separator after first row (header)
        if i == 0:
            markdown_lines.append("| " + " | ".join(["---"] * max_cols) + " |")

    markdown_text = "\n".join(markdown_lines)

    # Create JSONL record
    file_hash = compute_file_hash(source_path)
    mime_type = get_mime_type(source_path)
    doc_id = file_hash[:16]

    schema_version = config.get("schema", {}).get("version", "2.2.0")
    processed_at = datetime.utcnow().isoformat() + "Z"

    chunk_tokens = len(markdown_text.split())

    record = {
        "id": f"{doc_id}_{sheet_name}_{chunk_info['row_span'][0]:04d}",
        "doc_id": doc_id,
        "chunk_index": chunk_info['row_span'][0],
        "text": markdown_text,
        "attrs": {
            "page_span": None,
            "sections": [],
            "table": True,
            "sheet": sheet_name,
            "row_span": chunk_info['row_span'],
            "n_rows": len(rows),
            "n_cols": max_cols,
            "has_header": chunk_info['has_header'],
            "has_blank_boundary": chunk_info['has_blank_boundary'],
            "token_count": chunk_tokens
        },
        "source": {
            "file_path": str(source_path.resolve()),
            "file_name": source_path.name,
            "file_type": source_path.suffix.lstrip('.'),
            "mime_type": mime_type
        },
        "metadata": {
            "schema_version": schema_version,
            "file_type": "xlsx",
            "mime_type": mime_type,
            "hash_input_sha256": file_hash,
            "processor": "openpyxl",
            "processor_version": "3.1.0",
            "processed_at": processed_at,
            "batch_id": batch_id,
            "processing_duration_ms": None,
            "confidence_score": 1.0,
            "warnings": []
        }
    }

    return markdown_text, [record]
